import asyncio
import aiogram.utils.exceptions
from openpyxl import load_workbook
import re
from aiogram import Bot, Dispatcher, executor, types

API_TOKEN = '6199212552:AAGL0z0iUrESTvxPNo2KrkgAeYLe5hDb9-s'

bot = Bot(token=API_TOKEN)
dp = Dispatcher(bot)


@dp.message_handler(commands=['start'])
async def start(message: types.Message):
    if message.from_user.last_name is None:
        name = f"{message.from_user.first_name}"
    else:
        name = f"{message.from_user.first_name} {message.from_user.last_name}"
    mess = f'Привіт, {name}'
    await message.answer(text=mess)
    await message.delete()


@dp.message_handler()
async def send_user_id(message: types.Message):
    if message.text.isdigit():
        data1 = await contacts('contact.xlsx', message.text)
        data2 = await prowaiders('prowaiders.xlsx', message.text)
        try:
            if data1 and data2:
                await message.answer(data1 + '\n' + data2, parse_mode='HTML')
            elif data1 and not data2:
                await message.answer(data1)
            else:
                await message.answer(data2, parse_mode='HTML')
        except aiogram.utils.exceptions.MessageTextIsEmpty:
            await message.answer('Немає данних про провайдера')

        # await message.answer(data1 + data2, parse_mode='HTML')


async def contacts(file_path, tt):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    for index, row in enumerate(sheet.iter_rows(values_only=True)):
        if index != 0 and row[1].split()[0] == str(tt):
            address_match = re.search(r"\(([^)]+)\)", row[1])  # Знаходимо адресу у круглих дужках
            address = address_match.group(1) if address_match else ''  # Витягуємо адресу з результату пошуку
            text = f"Регіон: {row[0]}\n" \
                   f"ТТ: {row[1].split()[0]}\n" \
                   f"Адреса: {address}\n" \
                   f"ТМ: {row[2]}\n" \
                   f"Тел.: <a>{row[3] if row[3] else ''}{', ' if row[3] and row[4] else ''}{row[4] if row[4] else ''}</a>\n" \
                   f"Графік: {row[5]}\n"
            return text


async def prowaiders(file_path, tt):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    for index, row in enumerate(sheet.iter_rows(values_only=True)):
        if index != 0 and row[0] == int(tt):
            text = f"<u><b>Інформація про провайдера</b></u>\n" \
                   f"Провайдер: {row[2] if row[2] else '-'}\n" \
                   f"Тел.: {row[3] if row[3] else '-'}\n" \
                   f"Договір на: <u>{row[4] if row[4] else '-'}</u>\n" \
                   f"№ Договору: <u>{row[5] if row[5] else '-'}</u>\n" \
                   f"IP: <b>{row[8] if row[8] else '-'}:8080</b>\n" \
                   f"Логін: {row[6] if row[6] else ''}\n" \
                   f"Пароль: {row[7] if row[7] else '-'}\n" \
                   f"Доп. інф.: {row[11] if row[11] else '-'}"
            return text


if __name__ == '__main__':
    executor.start_polling(dp, skip_updates=True)
